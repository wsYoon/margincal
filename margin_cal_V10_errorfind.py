import sys
import pandas as pd
import numpy as np
import requests
import json
import openpyxl
import xlrd
import xlwt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
import csv
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import *
from PyQt5 import uic
import shutil
import math
import os

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("margin_cal_V10.ui")[0]

#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("스낵365 RPA")
        self.btn_cal.clicked.connect(self.button1Function)          #마진계산 버튼
        self.btn_stockorder.clicked.connect(self.button2Function)   #위험수량 발주 버튼
        self.pushButton_save.clicked.connect(self.marketsave)       #마켓수수료 저장 버튼

        #마켓수수료 테이블 세팅
        self.table_market.setRowCount(20)
        self.table_market.setColumnCount(6)
        self.table_market.setColumnWidth(1,60)
        self.table_market.setColumnWidth(2,50)
        
        #마켓수수료 저장된 csv 파일 load
        f = open('마켓수수료.csv','r')
        rdr = csv.reader(f)
        i=0
        for line in rdr:
            self.table_market.setItem(i,0,QTableWidgetItem(line[0]))
            self.table_market.setItem(i,1,QTableWidgetItem(line[1]))
            self.table_market.setItem(i,2,QTableWidgetItem('%'))
            i=i+1
        f.close()
        self.pushButton_snack.clicked.connect(self.fileopen_snack)
        self.pushButton_order.clicked.connect(self.fileopen_order)
        self.pushButton_stock.clicked.connect(self.fileopen_stock)
        

    def button1Function(self) :
        #입력된 마켓 별 수수료 테이블 가져오기
        if os.path.isfile("errorfind.csv"):
            os.remove("errorfind.csv")

        ###########################################errorfind : 시작
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['프로그램 시작'])
        f.close()

        rowcount = self.table_market.rowCount()
        market = []
        market_fee = []
        for i in range(rowcount):
            if(self.table_market.item(i,0).text()=='끝'):
                break
            market.append(self.table_market.item(i,0).text())
            market_fee.append(float(self.table_market.item(i,1).text()))


        #230306 묶음상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
        bundle_item = {}
        #230314 낱개상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리, 묶음과 낱개 모두 합친 총합 딕셔너리
        one_item = {}
        total_item = {}

        #snacklist = pd.read_excel('D:\SW_job\job1_margin\sourcefile\상품목록.xlsx',sheet_name='Sheet1',engine='openpyxl')
        #snacklist = pd.read_excel(filename_snack[0],sheet_name='Sheet1',engine='openpyxl')
        #orderlist = pd.read_excel(filename_order[0],sheet_name='WorkSheet',engine='openpyxl')
        snacklist = pd.read_excel(filename_snack[0],engine='openpyxl')
        #선물세트 = pd.read_excel(filename_snack[0],sheet_name='선물세트',engine='openpyxl') #선물세트 불러오기
        orderlist = pd.read_excel(filename_order[0],engine='xlrd')
        paydeliverylist = pd.read_excel("230405_전마켓 유료배송 취합.xlsx",engine='openpyxl')
        orderlist = orderlist[:-1] #마지막행(합계 행) 삭제
        orderlist2 = orderlist.copy() # 송장번호 순으로 정렬할 별도의 데이터프레임 생성
        orderlist = orderlist.sort_values(by='관리번호') #관리번호에 따른 오름차순 정렬
        orderlist = orderlist.reset_index(drop=True)
        orderlist2 = orderlist2.sort_values(by='송장번호') #송장번호에 따른 오름차순 정렬
        orderlist2 = orderlist2.reset_index(drop=True)
        
        #묶음상품에서 판매가가 적혀있는 상품을 최상단으로 이동
        for i in range(len(orderlist['관리번호'])-2):
            cnt = 0
            if (orderlist['관리번호'][i] == orderlist['관리번호'][i+1]):
                orderlist_temp = []
                i_remember = i
                #orderlist_temp.append(orderlist.iloc[i,:])
                #i += 1
                cnt = 0
                while (orderlist['관리번호'][i] == orderlist['관리번호'][i+1]):
                    orderlist_temp.append(orderlist.iloc[i,:])
                    i += 1
                    cnt += 1
                    if(i==len(orderlist['관리번호'])-1):
                        break
                orderlist_temp.append(orderlist.iloc[i,:])
                orderlist_temp = pd.DataFrame(orderlist_temp)
                orderlist_temp = orderlist_temp.sort_values(by='판매가',ascending=False)
                orderlist.iloc[i_remember:i_remember+cnt+1,:] = orderlist_temp
        #오름차순 정렬된 데이터로 주문건 파일 복사하기
        filename_order_temp = filename_order[0].replace('.xls','')
        with pd.ExcelWriter(filename_order_temp + '_계산결과.xlsx') as writer:
            orderlist.to_excel(writer,sheet_name="WorkSheet",index=False)


        ##########################################errorfind : 계산결과 파일 생성
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['계산결과 파일 생성'])
        f.close()


        #파일 수정을 위해서 열기
        wb=openpyxl.load_workbook(filename_order_temp + '_계산결과.xlsx')
        #wb_copy = openpyxl.Workbook()
        sheet = wb.active
        
        # 총판매금액, 총 마진금액
        total_sell = 0
        total_margin = 0

        # 관리번호가 같은 경우 상품 정산금액 더하는 변수
        cost_origin_sum = 0

        for i in range(len(orderlist['관리번호'])):
            print(i,orderlist['상품명'][i])

            if((snacklist['상품명']==orderlist['상품명'][i]).any()==False): #상품목록에 없는 상품 주문건일 때
                sheet.cell(row=i+2, column=26).value = 0
                sheet.cell(row=i+2, column=27).value = 0
                continue
            if(pd.isna(snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['가격'].values[0])): #가격이 없는 상품 주문건일 때
                sheet.cell(row=i+2, column=26).value = 0
                sheet.cell(row=i+2, column=27).value = 0
                continue
            
            #수수료 설정
            fee = market_fee[market.index(orderlist['판매처'][i])]
            
            #정산금액
            cost_cal = orderlist['판매가'][i] * (1-fee*0.01)
            #마켓수수료
            cost_fee = orderlist['판매가'][i] * (fee*0.01)

            #상품 수량에 따른 배송비, 박스비 조건 수량
            cnt_condition1 = snacklist['배송비조건1'].loc[snacklist['상품명'] == orderlist['상품명'][i]]
            cnt_condition2 = snacklist['배송비조건2'].loc[snacklist['상품명'] == orderlist['상품명'][i]]
            cnt_condition3 = snacklist['배송비조건3'].loc[snacklist['상품명'] == orderlist['상품명'][i]]
            cnt_condition4 = snacklist['배송비조건4'].loc[snacklist['상품명'] == orderlist['상품명'][i]]

            #박스비, 배송비
            if(orderlist['상품수량'][i] <= int(cnt_condition1)):
                box = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['박스비1']
                parcel = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['배송비1']
                
            elif(orderlist['상품수량'][i] <= int(cnt_condition2)):
                box = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['박스비2']
                parcel = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['배송비2']
            elif(orderlist['상품수량'][i] <= int(cnt_condition3)):
                box = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['박스비3']
                parcel = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['배송비3']
            elif(orderlist['상품수량'][i] <= int(cnt_condition4)):
                box = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['박스비4']
                parcel = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['배송비4']
            else:
                box = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['박스비5']
                parcel = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['배송비5']
                
            #원가*상품수량 (상품목록.xlsx에서 가져오기)
            cost_origin = snacklist.loc[snacklist['상품명'] == orderlist['상품명'][i]]['가격'].values[0] * orderlist['상품수량'][i]

            #유료배송비
            paydelivery = 0
            paydeliverylist.iloc[:,1]=paydeliverylist.iloc[:,1].astype(str)
            #print(orderlist['판매처 상품코드'][i])
            if str(orderlist['판매처 상품코드'][i]) in paydeliverylist.iloc[:,1].tolist():
                paydelivery =  paydeliverylist.loc[paydeliverylist['판매처상품코드'] == orderlist['판매처 상품코드'][i]]['배송비'].values[0]
                if isinstance(paydelivery, str):
                    paydelivery = int(paydelivery.replace(',',''))

            #원가, 배송비, 박스비를 제외한 마진금액
            if(orderlist['판매처'][i] == '오프라인'): #오프라인은 별도 배송비, 박스비 없이 수수료만으로 일괄계산
                cost_margin = cost_cal+paydelivery-cost_origin
            else:
                cost_margin = cost_cal+paydelivery-cost_origin-box.values[0]-parcel.values[0]
            
            #묶음상품이어서 판매가가 0원으로 기록된 경우는 margin_per 계산 제외
            if(orderlist['판매가'][i] != 0):
                margin_per = cost_margin / orderlist['판매가'][i] * 100
            else:
                margin_per = 0


            #관리번호가 같은 경우(묶음상품)
            if(i==0 and orderlist['관리번호'][i] != orderlist['관리번호'][i+1]): #첫번째 행이면서 묶음상품이 아닐 때
                sheet.cell(row=i+2, column=23).value = cost_cal
                sheet.cell(row=i+2, column=24).value = cost_fee
                sheet.cell(row=i+2, column=25).value = paydelivery
                sheet.cell(row=i+2, column=26).value = cost_margin
                sheet.cell(row=i+2, column=27).value = margin_per
            elif(i==len(orderlist['관리번호'])-1 and (orderlist['관리번호'][i] == orderlist['관리번호'][i-1])): #마지막행이면서 묶음상품일 때
                cost_origin_sum = cost_origin + cost_origin_sum
                cost_margin = cost_cal_memory+paydelivery-cost_origin_sum-box_remember-parcel_remember
                margin_per = cost_margin / orderlist['판매가'][i_memory] * 100
                sheet.cell(row=i+2, column=23).value = 0
                sheet.cell(row=i+2, column=24).value = 0
                sheet.cell(row=i+2, column=25).value = 0
                sheet.cell(row=i+2, column=26).value = 0
                sheet.cell(row=i+2, column=27).value = 0
                sheet.cell(row=i_memory+2, column=25).value = paydelivery
                sheet.cell(row=i_memory+2, column=26).value = cost_margin
                sheet.cell(row=i_memory+2, column=27).value = margin_per
                cost_origin_sum = 0
            elif(i==len(orderlist['관리번호'])-1): #주문목록의 마지막 행일 때
                sheet.cell(row=i+2, column=23).value = cost_cal
                sheet.cell(row=i+2, column=24).value = cost_fee
                sheet.cell(row=i+2, column=25).value = paydelivery
                sheet.cell(row=i+2, column=26).value = cost_margin
                sheet.cell(row=i+2, column=27).value = margin_per
            elif (orderlist['관리번호'][i] == orderlist['관리번호'][i+1]): #묶음상품의 마지막행이 아닐 때
                cost_origin_sum = cost_origin + cost_origin_sum
                sheet.cell(row=i+2, column=23).value = 0
                sheet.cell(row=i+2, column=24).value = 0
                sheet.cell(row=i+2, column=25).value = 0
                sheet.cell(row=i+2, column=26).value = 0
                sheet.cell(row=i+2, column=27).value = 0
                if(i==0):
                    i_memory = i #묶음상품 첫번째 행 저장
                    cost_cal_memory=cost_cal
                    box_remember = box.values[0]
                    parcel_remember = parcel.values[0]
                    sheet.cell(row=i+2, column=23).value = cost_cal
                    sheet.cell(row=i+2, column=24).value = cost_fee
                elif(orderlist['관리번호'][i] != orderlist['관리번호'][i-1]): #묶음상품의 첫번째 행일 때
                    i_memory = i #묶음상품 첫번째 행 저장
                    cost_cal_memory=cost_cal
                    box_remember = box.values[0]
                    parcel_remember = parcel.values[0]
                    sheet.cell(row=i+2, column=23).value = cost_cal
                    sheet.cell(row=i+2, column=24).value = cost_fee
            elif((orderlist['관리번호'][i] == orderlist['관리번호'][i-1]) and (orderlist['관리번호'][i] != orderlist['관리번호'][i+1])):#묶음상품의 마지막행일 때
                cost_origin_sum = cost_origin + cost_origin_sum
                cost_margin = cost_cal_memory+paydelivery-cost_origin_sum-box_remember-parcel_remember
                margin_per = cost_margin / orderlist['판매가'][i_memory] * 100
                sheet.cell(row=i+2, column=23).value = 0
                sheet.cell(row=i+2, column=24).value = 0
                sheet.cell(row=i+2, column=25).value = 0
                sheet.cell(row=i+2, column=26).value = 0
                sheet.cell(row=i+2, column=27).value = 0
                sheet.cell(row=i_memory+2, column=25).value = paydelivery
                sheet.cell(row=i_memory+2, column=26).value = cost_margin
                sheet.cell(row=i_memory+2, column=27).value = margin_per
                cost_origin_sum = 0
            else: # 묶음상품이 아닐 때
                sheet.cell(row=i+2, column=23).value = cost_cal
                sheet.cell(row=i+2, column=24).value = cost_fee
                sheet.cell(row=i+2, column=25).value = paydelivery
                sheet.cell(row=i+2, column=26).value = cost_margin
                sheet.cell(row=i+2, column=27).value = margin_per
            
            ##########################################errorfind : 마진계산 상품명
            f = open("errorfind.csv",'a',newline='')
            errorfind = csv.writer(f)
            errorfind.writerow(['',orderlist['관리번호'][i]])
            f.close()
        
        ##########################################errorfind : 마진계산 종료
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['마진계산 종료'])
        f.close()


        box_20EAover = pd.DataFrame([["상품명","상품수량","송장수량"]]) # 동일 구성이 20박스가 넘는 상품
        box_20EAover = box_20EAover.rename(columns=box_20EAover.iloc[0])
        box_20EAover = box_20EAover.drop(box_20EAover.index[0])
        
        # 송장번호에 따라서 합포상품에 들어가는 상품 수 세기
        for i in range(len(orderlist2['송장번호'])):
            #송장번호가 같은 경우(합포상품)
            if(i==0 and orderlist2['송장번호'][i] != orderlist2['송장번호'][i+1]): #첫번째 행이면서 묶음상품이 아닐 때
                #230314 낱개상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in one_item:
                    one_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    one_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
                #230413 동일 구성이 20박스가 넘는 상품(묶음 상품 아닌 경우만)
                if len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i])])==0:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)
                elif len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i]) & (box_20EAover['상품수량']==orderlist2['상품수량'][i])])>0:
                    box_20EAover['송장수량'].loc[(box_20EAover['상품명'] == orderlist2['상품명'][i]) & (box_20EAover['상품수량'] == orderlist2['상품수량'][i])] += 1
                else:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)
            elif(i==len(orderlist2['송장번호'])-1 and (orderlist2['송장번호'][i] == orderlist2['송장번호'][i-1])): #마지막행이면서 묶음상품일 때
                #230306 묶음상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in bundle_item:
                    bundle_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    bundle_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
            elif(i==len(orderlist2['송장번호'])-1): #주문목록의 마지막 행일 때
                #230314 낱개상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in one_item:
                    one_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    one_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
                #230413 동일 구성이 20박스가 넘는 상품(묶음 상품 아닌 경우만)
                if len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i])])==0:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)
                elif len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i]) & (box_20EAover['상품수량']==orderlist2['상품수량'][i])])>0:
                    box_20EAover['송장수량'].loc[(box_20EAover['상품명'] == orderlist2['상품명'][i]) & (box_20EAover['상품수량'] == orderlist2['상품수량'][i])] += 1
                else:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)
            elif (orderlist2['송장번호'][i] == orderlist2['송장번호'][i+1]): #묶음상품의 마지막행이 아닐 때
                #230306 묶음상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in bundle_item:
                    bundle_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    bundle_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
            elif((orderlist2['송장번호'][i] == orderlist2['송장번호'][i-1]) and (orderlist2['송장번호'][i] != orderlist2['송장번호'][i+1])):#묶음상품의 마지막행일 때
                #230306 묶음상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in bundle_item:
                    bundle_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    bundle_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
            else: # 묶음상품이 아닐 때
                #230314 낱개상품에 들어가는 상품들의 낱개 개수의 총합을 계산하는 딕셔너리
                if orderlist2['상품명'][i] in one_item:
                    one_item[orderlist2['상품명'][i]] += orderlist2['상품수량'][i]
                else:
                    one_item[orderlist2['상품명'][i]] = orderlist2['상품수량'][i]
                #230413 동일 구성이 20박스가 넘는 상품(묶음 상품 아닌 경우만)
                if len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i])])==0:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)
                elif len(box_20EAover[(box_20EAover['상품명']==orderlist2['상품명'][i]) & (box_20EAover['상품수량']==orderlist2['상품수량'][i])])>0:
                    box_20EAover['송장수량'].loc[(box_20EAover['상품명'] == orderlist2['상품명'][i]) & (box_20EAover['상품수량'] == orderlist2['상품수량'][i])] += 1
                else:
                    box_20EAover = box_20EAover.append({'상품명' : orderlist2['상품명'][i], '상품수량' : orderlist2['상품수량'][i], '송장수량' : 1}, ignore_index=True)

        ##########################################errorfind : 합포시트 상품 수량 세기 종료
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['합포시트 상품 수량 세기 종료'])
        f.close()

        #230306 묶음상품에 들어가는 상품들의 낱개 개수의 총합을 기록하는 시트 추가
        wb.create_sheet('합포상품수량')
        ws2 = wb['합포상품수량']
        ws2.cell(row=1,column=1).value = '상품명'
        ws2.cell(row=1,column=2).value = '공급처'
        ws2.cell(row=1,column=3).value = '상품총수량'
        ws2.cell(row=1,column=4).value = '입수'
        ws2.cell(row=1,column=5).value = '박스수'
        ws2.cell(row=1,column=6).value = '낱개'
        ws2.cell(row=1,column=7).value = '위치'
        i_bundle = 0
        for key, value in bundle_item.items():
            print(key, value)
            i_bundle += 1
            ws2.cell(row=i_bundle+1,column=1).value = key
            ws2.cell(row=i_bundle+1,column=2).value = str(snacklist.loc[snacklist['상품명'] == key,'공급처'].values.item())
            ws2.cell(row=i_bundle+1,column=3).value = value
            ws2.cell(row=i_bundle+1,column=4).value = int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws2.cell(row=i_bundle+1,column=5).value = value //  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws2.cell(row=i_bundle+1,column=6).value = value %  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws2.cell(row=i_bundle+1,column=7).value = str(snacklist.loc[snacklist['상품명'] == key,'위치'].values.item())

        #230314 낱개상품에 들어가는 상품들의 낱개 개수의 총합을 기록하는 시트 추가
        wb.create_sheet('낱개상품수량')
        ws3 = wb['낱개상품수량']
        ws3.cell(row=1,column=1).value = '상품명'
        ws3.cell(row=1,column=2).value = '공급처'
        ws3.cell(row=1,column=3).value = '상품총수량'
        ws3.cell(row=1,column=4).value = '입수'
        ws3.cell(row=1,column=5).value = '박스수'
        ws3.cell(row=1,column=6).value = '낱개'
        ws3.cell(row=1,column=7).value = '위치'
        i_one = 0
        for key, value in one_item.items():
            i_one += 1
            ws3.cell(row=i_one+1,column=1).value = key
            ws3.cell(row=i_one+1,column=2).value = str(snacklist.loc[snacklist['상품명'] == key,'공급처'].values.item())
            ws3.cell(row=i_one+1,column=3).value = value
            ws3.cell(row=i_one+1,column=4).value = int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws3.cell(row=i_one+1,column=5).value = value //  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws3.cell(row=i_one+1,column=6).value = value %  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws3.cell(row=i_one+1,column=7).value = str(snacklist.loc[snacklist['상품명'] == key,'위치'].values.item())

        #230314 토탈상품 낱개 개수의 총합을 기록하는 시트 추가
        for key in one_item.keys():
            if key not in bundle_item.keys():
                continue
            one_item[key] += bundle_item[key]
        for key in bundle_item.keys():
            if key not in one_item.keys():
                one_item[key] = bundle_item[key]
        wb.create_sheet('토탈상품수량')
        ws4 = wb['토탈상품수량']
        ws4.cell(row=1,column=1).value = '상품명'
        ws4.cell(row=1,column=2).value = '공급처'
        ws4.cell(row=1,column=3).value = '상품총수량'
        ws4.cell(row=1,column=4).value = '입수'
        ws4.cell(row=1,column=5).value = '박스수'
        ws4.cell(row=1,column=6).value = '낱개'
        ws4.cell(row=1,column=7).value = '위치'
        i_one = 0
        for key, value in one_item.items():
            i_one += 1
            ws4.cell(row=i_one+1,column=1).value = key
            ws4.cell(row=i_one+1,column=2).value = str(snacklist.loc[snacklist['상품명'] == key,'공급처'].values.item())
            ws4.cell(row=i_one+1,column=3).value = value
            ws4.cell(row=i_one+1,column=4).value = int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws4.cell(row=i_one+1,column=5).value = value //  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws4.cell(row=i_one+1,column=6).value = value %  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws4.cell(row=i_one+1,column=7).value = str(snacklist.loc[snacklist['상품명'] == key,'위치'].values.item())

        #230413 동일 구성이 20박스가 넘는 상품(묶음 상품 아닌 경우만)
        wb.create_sheet('20박스이상상품')
        ws5 = wb['20박스이상상품']
        ws5.cell(row=1,column=1).value = '상품명'
        ws5.cell(row=1,column=2).value = '상품수량'
        ws5.cell(row=1,column=3).value = '송장수량'
        ws5.cell(row=1,column=4).value = '위치'
        j=0
        for i in range(len(box_20EAover)):
            if(box_20EAover['송장수량'][i] >= 20):
                ws5.cell(row=j+2,column=1).value = box_20EAover['상품명'][i]
                ws5.cell(row=j+2,column=2).value = box_20EAover['상품수량'][i]
                ws5.cell(row=j+2,column=3).value = box_20EAover['송장수량'][i]
                ws5.cell(row=j+2,column=4).value = str(snacklist.loc[snacklist['상품명'] == box_20EAover['상품명'][i],'위치'].values.item())
                j+=1

        # 토탈상품 수량에서 '20박스이상상품'을 빼놓고 남은 수량만 별도 표기하는 시트
        for i in range(len(box_20EAover)):
            if(box_20EAover['송장수량'][i] >= 20):
                one_item[box_20EAover['상품명'][i]] -= box_20EAover['상품수량'][i]*box_20EAover['송장수량'][i]
        wb.create_sheet('그 외상품(20박스미만)')
        ws6 = wb['그 외상품(20박스미만)']
        ws6.cell(row=1,column=1).value = '상품명'
        ws6.cell(row=1,column=2).value = '공급처'
        ws6.cell(row=1,column=3).value = '상품총수량'
        ws6.cell(row=1,column=4).value = '입수'
        ws6.cell(row=1,column=5).value = '박스수'
        ws6.cell(row=1,column=6).value = '낱개'
        ws6.cell(row=1,column=7).value = '위치'
        i_one = 0
        for key, value in one_item.items():
            i_one += 1
            ws6.cell(row=i_one+1,column=1).value = key
            ws6.cell(row=i_one+1,column=2).value = str(snacklist.loc[snacklist['상품명'] == key,'공급처'].values.item())
            ws6.cell(row=i_one+1,column=3).value = value
            ws6.cell(row=i_one+1,column=4).value = int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws6.cell(row=i_one+1,column=5).value = value //  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws6.cell(row=i_one+1,column=6).value = value %  int(snacklist['입수'].loc[snacklist['상품명'] == key])
            ws6.cell(row=i_one+1,column=7).value = str(snacklist.loc[snacklist['상품명'] == key,'위치'].values.item())

        ##########################################errorfind : 상품수량 시트 생성 종료
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['상품수량 시트 생성 종료'])
        f.close()

        #230801 선물세트

        wb.save(filename_order_temp + '_계산결과.xlsx')

        
        #GUI에 총판매가, 마진, 마진율 표기하기 위한 계산
        orderlist = pd.read_excel(filename_order_temp + '_계산결과.xlsx',sheet_name='WorkSheet',engine='openpyxl')
        for i in range(len(orderlist['관리번호'])):
            total_sell = total_sell + orderlist['판매가'][i] + orderlist['유료배송비'][i] # 총 판매가
            total_margin = total_margin + orderlist['마진금액'][i] # 총 마진금액
        total_marginfee = total_margin/total_sell * 100
        self.lineEdit_sell.setText(format(total_sell,','))
        self.lineEdit_margin.setText(format(round(total_margin),','))
        self.lineEdit_marginfee.setText(str(round(total_marginfee,2)))
        sheet.cell(row=len(orderlist['관리번호'])+2, column=22).value = format(total_sell,',')
        sheet.cell(row=len(orderlist['관리번호'])+2, column=26).value = format(round(total_margin),',')
        sheet.cell(row=len(orderlist['관리번호'])+2, column=27).value = round(total_marginfee,2)

        wb.save(filename_order_temp + '_계산결과.xlsx')

        i_market = 0
        for txt_market in market:
            market_sell = 0
            market_margin =0
            for i in range(len(orderlist['관리번호'])):
                if (orderlist['판매처'][i] == txt_market):
                    market_sell = market_sell + orderlist['판매가'][i]
                    market_margin = market_margin + orderlist['마진금액'][i]
            if(market_margin==0):
                i_market = i_market+1
                continue
            if market_margin<0: # 자체 배송의 경우 판매가 기록을 하지 않아서 마진이 (-)가 되고 아래 수식에서 에러가 발생하므로 예외처리
                market_marginfee = 0
            else:
                market_marginfee = market_margin/market_sell*100
            self.table_market.setItem(i_market,3,QTableWidgetItem(format(market_sell,',')))
            self.table_market.setItem(i_market,4,QTableWidgetItem(format(round(market_margin),',')))
            self.table_market.setItem(i_market,5,QTableWidgetItem(str(round(market_marginfee,2))+'%'))
            i_market = i_market+1

        ##########################################errorfind : GUI에 마진 표기하고 계산결과파일 저장
        f = open("errorfind.csv",'a',newline='')
        errorfind = csv.writer(f)
        errorfind.writerow(['GUI에 마진 표기하고 계산결과파일 저장'])
        f.close()    

    def button2Function(self):
        snacklist = pd.read_excel(filename_snack[0],engine='openpyxl')  #원가표
        stocklist = pd.read_excel(filename_stock[0],engine='openpyxl')      #현재고조회
        shutil.copy("발주서 자동완성 요청양식.xlsx", "발주서 자동완성 요청양식_결과.xlsx")
        wb=openpyxl.load_workbook('발주서 자동완성 요청양식_결과.xlsx')
        ws = wb["Worksheet"]

        for i in range(len(stocklist['상품명'])):
            if (stocklist['정상재고'][i]-stocklist['접수'][i])<stocklist['위험수량'][i]*1.2:  # 재고가 위험수량*1.2 이하일 때 발주서 공백에 표시
                if (stocklist['정상재고'][i]-stocklist['접수'][i])<stocklist['위험수량'][i]:  # 재고가 위험수량 이하일 때 발주서 작성
                    print(stocklist['상품명'][i])
                    N_onebox = snacklist['입수'].loc[snacklist['상품명'] == stocklist['상품명'][i]]
                    print(stocklist['위험수량'][i])
                    print(stocklist['정상재고'][i]-stocklist['접수'][i])
                    print(N_onebox)
                    print(int(N_onebox))
                    N_box = math.ceil((stocklist['위험수량'][i]-(stocklist['정상재고'][i]-stocklist['접수'][i]))/int(N_onebox))
                    sheetnames = wb.sheetnames
                    if stocklist['공급처'][i] in sheetnames:    # 해당 공급처 시트가 이미 있는 경우
                        row_cnt = 7 # 발주 상품을 적는 가장 상단이 7행에서 시작
                        sht = wb[stocklist['공급처'][i]]
                        while sht["B"+str(row_cnt)].value != None :
                            row_cnt += 1
                        sht.cell(row=row_cnt,column=2).value = stocklist['상품명'][i]
                        sht.cell(row=row_cnt,column=3).value = N_onebox.item()
                        sht.cell(row=row_cnt,column=4).value = N_box
                        sht.cell(row=row_cnt,column=6).value = snacklist['가격'].loc[snacklist['상품명'] == stocklist['상품명'][i]].item()
                    else:                                       # 해당 공급처 시트가 없어서 새로 생성하는 경우
                        # 새로운 워크북과 워크시트 생성
                        target = wb.copy_worksheet(ws)
                        target.title = stocklist['공급처'][i]
                        target.cell(row=4,column=2).value = stocklist['공급처'][i]
                        target.cell(row=7,column=2).value = stocklist['상품명'][i]
                        target.cell(row=7,column=3).value = N_onebox.item()
                        target.cell(row=7,column=4).value = N_box
                        target.cell(row=7,column=6).value = snacklist['가격'].loc[snacklist['상품명'] == stocklist['상품명'][i]].item()
                        wb.save('발주서 자동완성 요청양식_결과.xlsx')
                else:   #현재고가 위험수량보다 많고 위험수량*1.2보다 작을 때 발주 후보 리스트 작성
                    N_onebox = snacklist['입수'].loc[snacklist['상품명'] == stocklist['상품명'][i]]
                    N_box = math.ceil((stocklist['위험수량'][i]*1.2-(stocklist['정상재고'][i]-stocklist['접수'][i]))/int(N_onebox))
                    sheetnames = wb.sheetnames
                    if stocklist['공급처'][i] in sheetnames:    # 해당 공급처 시트가 이미 있는 경우
                        row_cnt2 = 7 # 발주 상품을 적는 가장 상단이 7행에서 시작
                        sht = wb[stocklist['공급처'][i]]
                        while sht["I"+str(row_cnt2)].value != None :
                            row_cnt2 += 1
                        sht.cell(row=row_cnt2,column=9).value = stocklist['상품명'][i]
                        sht.cell(row=row_cnt2,column=10).value = N_onebox.item()
                        sht.cell(row=row_cnt2,column=11).value = N_box
                        sht.cell(row=row_cnt2,column=13).value = snacklist['가격'].loc[snacklist['상품명'] == stocklist['상품명'][i]].item()
                    else:                                       # 해당 공급처 시트가 없어서 새로 생성하는 경우
                        # 새로운 워크북과 워크시트 생성
                        target = wb.copy_worksheet(ws)
                        target.title = stocklist['공급처'][i]
                        target.cell(row=4,column=2).value = stocklist['공급처'][i]
                        target.cell(row=7,column=9).value = stocklist['상품명'][i]
                        target.cell(row=7,column=10).value = N_onebox.item()
                        target.cell(row=7,column=11).value = N_box
                        target.cell(row=7,column=13).value = snacklist['가격'].loc[snacklist['상품명'] == stocklist['상품명'][i]].item()
                        wb.save('발주서 자동완성 요청양식_결과.xlsx')                    
                #print(N_box)
        wb.save('발주서 자동완성 요청양식_결과.xlsx')
        wb.close()

    def marketsave(self):
        market = []
        market_fee = []
        lines=[]
        #i=0
        for i in range(20):
            if(self.table_market.item(i,0)==None):
                line=[]
                line.append('')
                line.append('')
                lines.append(line)
            else:
                line=[]
                print(self.table_market.item(i,0).text())
                item1 = self.table_market.item(i,0).text()
                item2 = self.table_market.item(i,1).text()
                line.append(item1)
                line.append(item2)
                lines.append(line)
        line=[]
        line.append('끝')
        line.append('')
        lines.append(line)
        f = open('마켓수수료.csv','w',newline='')
        wr = csv.writer(f)
        wr.writerows(lines)
        f.close()

    def fileopen_snack(self):
        global filename_snack
        filename_snack = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_snack.setText(filename_snack[0])

    def fileopen_order(self):
        global filename_order
        filename_order = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_order.setText(filename_order[0])

    def fileopen_stock(self):
        global filename_stock
        filename_stock = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_stock.setText(filename_stock[0])

if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 

    #프로그램 화면을 보여주는 코드
    #myWindow.show()
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()